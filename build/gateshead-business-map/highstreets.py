"""High-street analysis for the Gateshead Business Map.

Adds, per Local Plan centre (see centres.py):
  * start-up activity (incorporations of currently-live companies, by quarter)
  * openings & closures from month-on-month Companies House snapshot diffs
    (persistent register in data/company_register.json — the free CH product
    is live-only, so disappearance = dissolved / moved / de-registered)
  * composition of FSA-rated premises (trading ground truth) + gap vs the
    all-centre average
  * independents vs chains (heuristic: national brand list + same name at
    3+ Gateshead premises — an analytical judgement, see assurance note)
  * walk-in catchment context (population + income deprivation, IMD 2025)
  * ONS/BT retail footfall context (North East region by site type — the
    dataset has NO per-town data; labelled as regional context only)
"""
from __future__ import annotations
import datetime as dt
import io
import json
import math
import re
from collections import Counter, defaultdict
from pathlib import Path

import config as C

DATA_DIR = Path(__file__).resolve().parent / "data"
REGISTER_PATH = DATA_DIR / "company_register.json"
CATCHMENT_M = 800.0          # walk-in catchment radius around centre centroid
FOOTFALL_PAGE = ("https://www.ons.gov.uk/economy/economicoutputandproductivity/"
                 "output/datasets/ukretailfootfall")

# --------------------------------------------------------------------------- #
# chains vs independents — heuristic brand list (national multiples commonly
# present on NE high streets). Matching is on normalised name prefix.
# --------------------------------------------------------------------------- #
BRANDS = [
    "greggs", "costa", "starbucks", "caffe nero", "subway", "mcdonald", "kfc",
    "burger king", "domino", "papa john", "pizza hut", "german doner",
    "tesco", "sainsbury", "asda", "morrisons", "aldi", "lidl", "iceland",
    "co-op", "coop", "co op", "farmfoods", "heron foods", "spar", "premier",
    "one stop", "nisa", "londis", "costcutter", "best one", "bestone",
    "boots", "superdrug", "savers", "bodycare", "specsavers", "vision express",
    "b&m", "home bargains", "poundland", "poundstretcher", "the works",
    "card factory", "wilko", "argos", "shoe zone", "sports direct", "jd sports",
    "peacocks", "primark", "new look", "next", "matalan", "tk maxx",
    "wh smith", "whsmith", "ryman", "post office", "timpson", "max spielmann",
    "ladbrokes", "william hill", "betfred", "coral", "paddy power", "bet365",
    "cash converters", "cex", "cash generator", "ramsdens", "h&t pawnbrokers",
    "lloyds pharmacy", "well pharmacy", "rowlands pharmacy", "boots pharmacy",
    "santander", "barclays", "halifax", "lloyds bank", "natwest", "hsbc",
    "tsb", "nationwide", "virgin money", "newcastle building society",
    "vodafone", "o2", "ee ", "three ", "screwfix", "toolstation", "howdens",
    "pets at home", "petshop", "vets4pets", "dominos", "krispy kreme",
    "taco bell", "nando", "wetherspoon", "marston", "stonegate",
]
_norm_re = re.compile(r"[^a-z0-9& ]+")

def _norm_name(n):
    n = _norm_re.sub(" ", (n or "").lower())
    n = re.sub(r"\b(ltd|limited|plc|llp|uk|the)\b", " ", n)
    return re.sub(r"\s+", " ", n).strip()

def flag_chains(fsa_pts):
    """Adds ch=1 (chain) / 0 to FSA points. Returns (n_chain, n_flagged_multi)."""
    counts = Counter(_norm_name(p["n"]) for p in fsa_pts if p.get("n"))
    multi = {n for n, k in counts.items() if n and k >= 3}
    n_chain = 0
    for p in fsa_pts:
        n = _norm_name(p.get("n"))
        is_brand = any(n.startswith(b) or (" " + b + " ") in (" " + n + " ") for b in BRANDS)
        p["ch"] = 1 if (is_brand or n in multi) else 0
        n_chain += p["ch"]
    return n_chain, len(multi)

# --------------------------------------------------------------------------- #
# FSA composition groups (trading ground truth for high-street mix)
# --------------------------------------------------------------------------- #
FSA_GROUPS = {
    "Restaurant/Cafe/Canteen": "Cafes & restaurants",
    "Takeaway/sandwich shop": "Takeaways",
    "Pub/bar/nightclub": "Pubs & bars",
    "Retailers - supermarkets/hypermarkets": "Supermarkets",
    "Retailers - other": "Food retail (other)",
    "Hotel/bed & breakfast/guest house": "Accommodation",
    "Mobile caterer": "Mobile caterers",
    "School/college/university": "Institutions",
    "Hospitals/Childcare/Caring Premises": "Institutions",
    "Manufacturers/packers": "Producers & distribution",
    "Importers/Exporters": "Producers & distribution",
    "Distributors/Transporters": "Producers & distribution",
    "Farmers/growers": "Producers & distribution",
    "Other catering premises": "Other catering",
}
GROUP_ORDER = ["Cafes & restaurants", "Takeaways", "Pubs & bars", "Supermarkets",
               "Food retail (other)", "Accommodation", "Other catering",
               "Institutions", "Producers & distribution", "Mobile caterers"]

def fsa_group(ty):
    return FSA_GROUPS.get(ty, "Other catering")

# --------------------------------------------------------------------------- #
# openings & closures register (month-on-month CH snapshot diff)
# --------------------------------------------------------------------------- #
def update_register(company_pts, snapshot, warn):
    """Diff current live companies against the persistent register.
    Returns (events, meta) and rewrites data/company_register.json."""
    DATA_DIR.mkdir(exist_ok=True)
    reg = {"snapshots": [], "companies": {}, "events": []}
    if REGISTER_PATH.exists():
        reg = json.loads(REGISTER_PATH.read_text(encoding="utf-8"))
    snaps = reg["snapshots"]
    comp = reg["companies"]
    seed = not snaps
    if snapshot in snaps:
        # idempotent within a month: refresh attributes, no new events
        si = snaps.index(snapshot)
    else:
        snaps.append(snapshot)
        si = len(snaps) - 1
    month = snapshot[:7]
    cur = {c["num"]: c for c in company_pts}
    new_events = []
    if not seed and si > 0:
        prev_last = si - 1
        for num, c in cur.items():
            if num not in comp:
                new_events.append({"m": month, "t": "new", "n": c["n"][:48],
                                   "sec": c["sec"], "ct": c.get("ct"), "lsoa": c["lsoa"]})
        for num, rec in comp.items():
            if rec["l"] == prev_last and num not in cur:
                new_events.append({"m": month, "t": "gone", "n": rec["n"],
                                   "sec": rec["s"], "ct": rec.get("c"), "lsoa": rec.get("z")})
    for num, c in cur.items():
        rec = comp.get(num)
        if rec is None:
            comp[num] = {"f": si, "l": si, "n": c["n"][:48], "s": c["sec"],
                         "c": c.get("ct"), "z": c["lsoa"]}
        else:
            rec["l"] = si
            rec["c"] = c.get("ct")
    # drop companies unseen for >30 snapshots to bound file size
    stale = [n for n, r in comp.items() if si - r["l"] > 30]
    for n in stale:
        del comp[n]
    reg["events"] = ([e for e in reg["events"] if e["m"] != month] + new_events)[-1500:]
    REGISTER_PATH.write_text(json.dumps(reg, separators=(",", ":")), encoding="utf-8")
    if seed:
        warn("openings/closures register seeded this run — change history builds from the next monthly refresh")
    return reg["events"], {"snapshots": snaps, "seeded": seed}

# --------------------------------------------------------------------------- #
# start-ups
# --------------------------------------------------------------------------- #
def _months_ago(im, now):
    y, m = int(im[:4]), int(im[5:7])
    return (now.year - y) * 12 + (now.month - m)

def startup_series(pts, now, quarters=12):
    """Quarterly incorporation counts (currently-live companies), oldest first."""
    qcounts = Counter()
    for c in pts:
        im = c.get("im")
        if not im:
            continue
        y, m = int(im[:4]), int(im[5:7])
        qcounts[(y, (m - 1) // 3 + 1)] += 1
    out = []
    y, q = now.year, (now.month - 1) // 3 + 1
    for _ in range(quarters):
        out.append({"q": f"{y} Q{q}", "n": qcounts.get((y, q), 0)})
        q -= 1
        if q == 0:
            y, q = y - 1, 4
    return out[::-1]

# --------------------------------------------------------------------------- #
# catchment context (population + income deprivation within CATCHMENT_M)
# --------------------------------------------------------------------------- #
def _feature_centroid(feat):
    g = feat.get("geometry") or {}
    coords = g.get("coordinates") or []
    ring = None
    if g.get("type") == "Polygon" and coords:
        ring = coords[0]
    elif g.get("type") == "MultiPolygon" and coords:
        ring = max((poly[0] for poly in coords), key=len)
    if not ring:
        return None
    lon = sum(p[0] for p in ring) / len(ring)
    lat = sum(p[1] for p in ring) / len(ring)
    return lat, lon

def catchments(centres, lsoa_feats, imd):
    """centre id -> {pop, inc_score (pop-weighted income-domain), dep30 share}."""
    cents = []
    for f in lsoa_feats:
        code = (f.get("properties") or {}).get("LSOA21CD")
        c = _feature_centroid(f)
        if code and c:
            cents.append((code, c[0], c[1]))
    mpd_lat = 111_320.0
    mpd_lon = 111_320.0 * math.cos(math.radians(54.95))
    out = {}
    for c in centres:
        pop = 0.0
        inc_acc = 0.0
        dep_pop = 0.0
        inc_w = 0.0
        for code, lat, lon in cents:
            d = math.hypot((lat - c["lat"]) * mpd_lat, (lon - c["lon"]) * mpd_lon)
            if d > CATCHMENT_M:
                continue
            rec = imd.get(code) or {}
            p = rec.get("pop") or 0
            pop += p
            dec = rec.get("decile")
            if dec is not None and dec <= 3:
                dep_pop += p
            inc = (rec.get("domains") or {}).get("income")
            if inc is not None and p:
                inc_acc += inc * p
                inc_w += p
        out[c["id"]] = {
            "pop": int(pop),
            "inc": round(inc_acc / inc_w, 3) if inc_w else None,
            "dep30": round(100 * dep_pop / pop, 1) if pop else None,
        }
    return out

# --------------------------------------------------------------------------- #
# centre rollups
# --------------------------------------------------------------------------- #
def centre_rollups(centres, company_pts, fsa_pts, imd, lsoa_feats, events, now, warn):
    cat = catchments(centres, lsoa_feats, imd)
    comp_by_ct = defaultdict(list)
    for c in company_pts:
        if c.get("ct"):
            comp_by_ct[c["ct"]].append(c)
    fsa_by_ct = defaultdict(list)
    for p in fsa_pts:
        if p.get("ct"):
            fsa_by_ct[p["ct"]].append(p)
    ev_by_ct = defaultdict(lambda: {"op": 0, "cl": 0})
    cutoff = (now.replace(day=1) - dt.timedelta(days=366)).strftime("%Y-%m")
    for e in events:
        if e.get("ct") and e["m"] >= cutoff:
            ev_by_ct[e["ct"]]["op" if e["t"] == "new" else "cl"] += 1

    rows = []
    for c in centres:
        cs = comp_by_ct.get(c["id"], [])
        fs = fsa_by_ct.get(c["id"], [])
        groups = Counter(fsa_group(p["ty"]) for p in fs)
        rated = [p for p in fs if p.get("ch") is not None]
        ind = (100 * sum(1 - p["ch"] for p in rated) / len(rated)) if len(rated) >= 5 else None
        tk = (100 * groups.get("Takeaways", 0) / len(fs)) if fs else None
        s12 = sum(1 for x in cs if x.get("im") and _months_ago(x["im"], now) < 12)
        s24 = sum(1 for x in cs if x.get("im") and _months_ago(x["im"], now) < 24)
        ct_cat = cat.get(c["id"], {})
        rows.append({
            "id": c["id"], "name": c["name"], "tier": c["tier"],
            "lat": c["lat"], "lon": c["lon"],
            "n": len(cs), "fsa": len(fs),
            "ind": None if ind is None else round(ind, 1),
            "tk": None if tk is None else round(tk, 1),
            "gr": {g: groups.get(g, 0) for g in GROUP_ORDER if groups.get(g, 0)},
            "s12": s12, "s24": s24,
            "sq": [x["n"] for x in startup_series(cs, now, 8)],
            "op": ev_by_ct[c["id"]]["op"], "cl": ev_by_ct[c["id"]]["cl"],
            "pop": ct_cat.get("pop"), "inc": ct_cat.get("inc"), "dep30": ct_cat.get("dep30"),
            "distress": sum(x["ds"] for x in cs),
        })
    # composition benchmark: average share by group across town+district+local centres
    hs_rows = [r for r in rows if r["tier"] != "destination" and r["fsa"] >= 5]
    avg = Counter()
    tot = 0
    for r in hs_rows:
        for g, k in r["gr"].items():
            avg[g] += k
        tot += r["fsa"]
    avg_share = {g: round(100 * avg[g] / tot, 1) for g in avg} if tot else {}
    return rows, avg_share

# --------------------------------------------------------------------------- #
# ONS/BT retail footfall — regional context (NO per-town data in the source)
# --------------------------------------------------------------------------- #
def load_footfall(session, warn):
    try:
        import openpyxl
        # ONS serves an HTML error to non-browser user agents on /file endpoints
        ua = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36"}
        page = session.get(FOOTFALL_PAGE, timeout=C.HTTP_TIMEOUT, headers=ua).text
        m = re.search(r'/file\?uri=[^"\']*ukretailfootfall[^"\']*\.xlsx', page)
        if not m:
            warn("footfall: no xlsx link found on ONS dataset page")
            return {"available": False}
        url = "https://www.ons.gov.uk" + m.group(0)
        r = session.get(url, timeout=120, headers=ua)
        if r.status_code != 200 or not r.content.startswith(b"PK"):
            # the /current/ alias 404s for this dataset — derive the edition path
            # from the filename date, e.g. ...dataset020726.xlsx -> /2july2026/
            fm = re.search(r"dataset(\d{2})(\d{2})(\d{2})\.xlsx", url)
            if fm:
                months = ["january", "february", "march", "april", "may", "june",
                          "july", "august", "september", "october", "november", "december"]
                edition = f"{int(fm.group(1))}{months[int(fm.group(2)) - 1]}20{fm.group(3)}"
                url = re.sub(r"/current/", f"/{edition}/", url)
                r = session.get(url, timeout=120, headers=ua)
        if r.status_code != 200 or not r.content.startswith(b"PK"):
            warn(f"footfall: download failed ({r.status_code})")
            return {"available": False}
        wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True)
        # Table 5: monthly by site type (UK); Table 6: monthly by region and site
        def sheet(name_part):
            for s in wb.sheetnames:
                if name_part in s:
                    return wb[s]
            return None
        ws5, ws6 = sheet("5.Monthly"), sheet("6.Monthly")
        if ws5 is None or ws6 is None:
            warn("footfall: expected sheets missing")
            return {"available": False}
        def parse(ws, want_cols):
            rows = ws.iter_rows(values_only=True)
            hdr = None
            for r in rows:
                if r[0] is not None and str(r[0]).strip().lower() == "month":
                    hdr = [str(c or "").strip() for c in r]
                    break
            idx = {}
            for label, needles in want_cols.items():
                for i, h in enumerate(hdr):
                    if all(n.lower() in h.lower() for n in needles):
                        idx[label] = i
                        break
            months, series = [], {k: [] for k in idx}
            for r in rows:
                if r[0] is None:
                    continue
                mth = str(r[0])[:7]
                months.append(mth)
                for k, i in idx.items():
                    try:
                        series[k].append(round(float(r[i]), 1))
                    except (TypeError, ValueError):
                        series[k].append(None)
            return months, series
        months, uk = parse(ws5, {
            "District or local centres": ["District"],
            "Retail parks": ["Retail Park"],
            "Town and city centres": ["Town and City"]})
        _, ne = parse(ws6, {
            "District or local centres": ["North East", "District"],
            "Retail parks": ["North East", "Retail Park"],
            "Town and city centres": ["North East", "Town and City"]})
        return {"available": True, "months": months, "uk": uk, "ne": ne,
                "src_url": url}
    except Exception as e:
        warn(f"footfall unavailable: {type(e).__name__}")
        return {"available": False}
