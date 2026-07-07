"""Parse the ONS Explore Local Statistics all-datasets workbook into the
reduced scorecard data object consumed by dashboard_src.html.

Per indicator (where Gateshead E08000037 has data): Gateshead full series,
best national benchmark (England > UK > GB > E&W) series, North East series,
NECA 7 + ONS economic statistical neighbour latest values, all-LAD latest
values (sorted, for rank/quartile/distribution), Gateshead 95% CI where the
table publishes one, source line, unit, period labels, polarity and domain.

Run via build.py; standalone: python parse_all.py [workbook.xlsx]
"""
import json
import re
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent

GH = "E08000037"
NE = "E12000001"
BENCH_PREF = ["E92000001", "K02000001", "K03000001", "K04000001"]
BENCH_NAME = {"E92000001": "England", "K02000001": "UK", "K03000001": "Great Britain", "K04000001": "England & Wales"}
NECA = ["E08000037", "E08000021", "E08000022", "E08000023", "E08000024", "E06000047", "E06000057"]
NEIGH = ["E08000036", "E08000013", "E08000016", "E06000047", "E08000014",
         "W06000022", "W06000014", "E08000024", "E06000052", "W06000004",
         "W06000009", "W06000008", "E06000057", "E06000046", "W06000003",
         "E06000005", "E08000017", "E08000027", "E06000066", "N09000010"]
LAD_RE = re.compile(r"^(E06|E07|E08|E09|W06|S12|N09)")

# polarity: +1 higher is better, -1 lower is better, 0 context/neutral
POL = {1:0, 3:0, 4:0, 5:0, 6:0, 7:0, 8:0,
       9:-1, 11:1, 13:-1, 14:-1, 15:1, 16:1, 17:-1, 18:-1,
       20:1, 21:1, 22:1, 23:1,
       28:0, 29:1, 30:-1, 31:1, 32:1, 33:1, 34:1, 35:1, 36:1,
       37:0, 38:1, 41:0, 42:-1, 43:1, 44:0, 45:0,
       46:1, 47:1, 48:1, 49:1, 51:1, 52:-1, 53:1, 54:1, 55:1,
       56:-1, 57:-1, 58:-1, 59:1, 60:1,
       64:-1, 65:1, 66:1, 67:-1, 68:-1, 69:-1, 70:-1, 71:-1, 72:-1,
       73:-1, 74:1, 75:-1, 78:1, 79:1, 80:-1, 81:1, 82:1, 83:1,
       84:0, 85:0, 86:-1, 87:0, 88:0, 89:0,
       90:1, 91:1, 92:1, 93:-1, 94:0, 95:1,
       99:1, 100:1, 101:1, 102:1, 103:1, 104:1, 105:1, 106:1, 107:1}

DOMAINS = [
    ("Population", [1, 3, 4, 5, 6, 7, 8]),
    ("Work & income", [11, 13, 9, 14, 16, 15, 17, 18]),
    ("Economy & productivity", [20, 21, 22, 23]),
    ("Business", [28, 29, 30, 31, 32, 33, 34, 35, 36]),
    ("Housing", [38, 41, 42, 43]),
    ("Education & skills", [53, 54, 55, 51, 48, 56, 57, 58, 46, 47, 59, 60, 49, 52]),
    ("Health & wellbeing", [78, 79, 64, 73, 65, 66, 67, 68, 69, 70, 71, 72, 74, 75, 80, 81, 82, 83]),
    ("Environment & energy", [86, 84, 85, 87, 88, 89]),
    ("Connectivity & transport", [90, 93, 91, 92, 95, 94]),
    ("Amenities & culture", [37, 99, 100, 101, 102, 103, 104, 105, 106, 107, 44, 45]),
]
INCLUDE = [t for _, ts in DOMAINS for t in ts]

MON = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

def plabel(h):
    """Column header -> short period label."""
    s = str(h)
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})(?:/P(\d+)([YMWD]))?", s)
    if not m:
        return s[:12]
    y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
    n = int(m.group(4)) if m.group(4) else None
    u = m.group(5)
    if u == "Y":
        if n == 1:
            return f"{y}" if mo == 1 else f"{y}/{str(y+1)[2:]}"
        end = y + n - 1 if mo == 1 else y + n
        return f"{y}–{str(end)[2:]}"
    if u == "M":
        return f"{MON[mo-1]} {y}"
    if (mo, d) == (1, 1):
        return f"{y}"
    if mo == 6 and d == 30:
        return f"mid-{y}"
    return f"{MON[mo-1]} {y}"

def num(v):
    if v is None:
        return None
    s = str(v).strip()
    if not s or s.startswith("["):
        return None
    try:
        return float(s)
    except ValueError:
        return None

def sig(v):
    if v is None:
        return None
    a = abs(v)
    if a >= 1000: return round(v)
    if a >= 100:  return round(v, 1)
    return round(v, 2)

def parse(workbook_path):
    wb = openpyxl.load_workbook(workbook_path, read_only=True)
    cover = [str(r[0]) for r in wb["Cover_sheet"].iter_rows(values_only=True) if r[0]]
    gen = next((l for l in cover if "generated" in l), "")
    gd = re.search(r"(\d{2})/(\d{2})/(\d{4})", gen)
    els_generated = f"{gd.group(3)}-{gd.group(2)}-{gd.group(1)}" if gd else ""

    names = {}
    ind = {}

    for t in INCLUDE:
        ws = wb[str(t)]
        rows = ws.iter_rows(values_only=True)
        title = unit = src = ""
        hdr = None
        for r in rows:
            c0 = str(r[0]).strip() if r[0] is not None else ""
            if c0.lower() == "area code":
                hdr = r
                break
            if not c0:
                continue
            low = c0.lower()
            if low.startswith("source"):
                src = c0
            elif low.startswith(("this worksheet", "some shorthand")):
                pass
            elif not title:
                title = re.sub(r"\s*\[note \d+\]\s*", "", c0).strip()
            elif not unit:
                unit = c0
        cols = [str(c) if c is not None else "" for c in hdr]
        has_measure = len(cols) > 2 and cols[2].strip().lower() == "measure"
        v0 = 3 if has_measure else 2
        ncols = len([c for c in cols if c])
        heads = cols[v0:ncols]
        labels = [plabel(h) for h in heads]

        um = re.search(r"\(([^)]*)\)\s*$", heads[0]) if heads else None
        unit_short = um.group(1) if um else ""
        if t == 42:
            unit_short = "ratio"   # ONS header says (%) but the measure is a ratio

        data = {}
        ci = {}
        for r in rows:
            code = str(r[0]) if r[0] is not None else ""
            if not code or code.lower() == "area code":
                continue
            vals = [num(v) for v in r[v0:ncols]]
            if has_measure:
                meas = str(r[2]).lower()
                if "lower conf" in meas or "upper conf" in meas:
                    if code == GH:
                        ci.setdefault(code, {})["lo" if "lower" in meas else "hi"] = vals
                    continue
            if code not in data:
                data[code] = vals
                if r[1] is not None:
                    names.setdefault(code, str(r[1]))

        if GH not in data or all(v is None for v in data[GH]):
            continue

        li = max(i for i, v in enumerate(data[GH]) if v is not None)
        bench_code = next((b for b in BENCH_PREF if b in data and data[b][li] is not None), None)
        lad_latest = {c: v[li] for c, v in data.items() if LAD_RE.match(c) and v[li] is not None}
        n = len(lad_latest)
        ghv = data[GH][li]
        rank_high = 1 + sum(1 for v in lad_latest.values() if v > ghv)

        nations = {c[0] for c in lad_latest}
        cov = ("UK" if nations >= {"E","W","S","N"} else
               "Great Britain" if nations >= {"E","W","S"} else
               "England & Wales" if nations >= {"E","W"} else
               "England" if nations == {"E"} else "+".join(sorted(nations)))

        keep = list(range(len(labels)))
        if len(labels) > 60:   # monthly series -> annual + latest
            keep = list(range(len(labels) - 1, -1, -12))[::-1]
        def pick(seq):
            return [sig(seq[i]) for i in keep] if seq else None

        gh_ci = None
        if GH in ci and "lo" in ci[GH] and "hi" in ci[GH]:
            lo, hi = ci[GH]["lo"][li], ci[GH]["hi"][li]
            if lo is not None and hi is not None:
                gh_ci = [sig(lo), sig(hi)]

        ind[t] = {
            "name": title, "unit": unit, "unitShort": unit_short,
            "src": src.replace("Source: ", ""), "pol": POL[t],
            "labels": [labels[i] for i in keep],
            "latest": labels[li],
            "gh": pick(data[GH]),
            "bench": pick(data[bench_code]) if bench_code else None,
            "benchCode": BENCH_NAME.get(bench_code),
            "ne": pick(data[NE]) if NE in data else None,
            "ghCI": gh_ci, "surveyCI": has_measure,
            "neca": {c: sig(data[c][li]) if c in data else None for c in NECA},
            "neigh": {c: sig(data[c][li]) if c in data else None for c in NEIGH},
            "ladSorted": sorted((sig(v) for v in lad_latest.values()), reverse=True),
            "n": n, "rankHigh": rank_high, "cov": cov,
        }

    return {
        "elsGenerated": els_generated,
        "gh": GH,
        "neca": NECA, "neigh": NEIGH,
        "names": {c: names.get(c, c) for c in sorted(set(NECA) | set(NEIGH))},
        "domains": [{"name": d, "tables": [t for t in ts if t in ind]} for d, ts in DOMAINS],
        "ind": ind,
    }

if __name__ == "__main__":
    import sys
    wbpath = Path(sys.argv[1]) if len(sys.argv) > 1 else BASE / "data" / "all-datasets.xlsx"
    out = parse(wbpath)
    dest = BASE / "data" / "scorecard_data.json"
    dest.write_text(json.dumps(out, separators=(",", ":"), ensure_ascii=False), encoding="utf-8")
    print(f"Indicators: {len(out['ind'])} -> {dest}")
